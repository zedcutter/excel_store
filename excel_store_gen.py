import sys, openpyxl, pprint
from openpyxl.utils import get_column_letter
# ---------------------------------------------------------------------------------------#
# EXCEL_STORE upload SQL and VIEW generator
#
#  MySQL command line: use <database> 
#                      source <generated script(s)>
#
#  Oracle sqlplus: set escape on
#                  set sqlblanklines on
#                  @<generated script(s)>
# 
#  P1 = excel file to process
#
# [P2] = sheet name pattern: 
#         "#ALL#" will process all sheets. This is the default if P2 is omitted.
#         Otherwise all sheeds will be processed where the name contains P2 (case sensitive). 
#         e.g. "PRODUCT" will process sheets "PRODUCT", "PRODUCTS Americas", "PRODUCT returns", etc.
#
# [P3] = generate view from header?
#         If not specified it's "Y"
#         in this case the view columns will be created from first row of the Excel
#         
#         if you specify anything else it will be treated as "N": it will
#         not generate a view
#
# Date          Edited              Change
# 2020-10-09    Zoltan Vago         Initial version. Limited to Excel columns A-Z.
# ---------------------------------------------------------------------------------------#

#String cleaners
def clean_str(p_str):
    return p_str.replace(' ','_').replace('.','_').replace('-','_').replace('/','_').replace('@','_').replace(':','_').replace('$','_').replace('&','_')

def clean_cellval(p_val):
    return p_val.replace("'","''").replace(';','\;\ ').replace('-','\-').replace('$','\$').replace('|','\|').replace('"','\"').replace('&',' and ')

#Check input params
param_cnt=len(sys.argv)-1

if param_cnt<1:
    print("Please specify at least P1 (excelfile)!")
    quit()
    
excelfile=str(sys.argv[1])
if param_cnt<2:
    wsheet_pattern='#ALL#'
else:
    wsheet_pattern=str(sys.argv[2])

if param_cnt<3:
    view_gen='Y' # header-based view
else:
    view_gen=str(sys.argv[3]) # whatever given, if Y >> view else no-view

#Process excel sheet(s)
print("Opening workbook... "+excelfile)
wb = openpyxl.load_workbook(excelfile, data_only=True)

# Get all sheet names
g_sheet=wb.sheetnames
for wsheet in g_sheet:
    
    if wsheet.find(wsheet_pattern)>-1 or wsheet_pattern=='#ALL#':
        print("Parsing sheet... "+wsheet)
        sheet=wb[wsheet]
        wsheet_display_name=clean_str(wsheet)
        # Open output file
        f_name='excel_store_ins_'+clean_str(excelfile)+'_'+wsheet_display_name+'.sql';
        f_view_name='excel_store_ins_'+clean_str(excelfile)+'_'+wsheet_display_name+'_v.sql';
        f_sql = open(f_name,'w')

        print('Processing rows...')
        f_sql.write("delete from excel_store where filename='"+excelfile+"';\n")
        for r in range(1, sheet.max_row + 1):
            # Each row in the spreadsheet has data 
            sql_expr="insert into excel_store (filename, rn, sheet, a,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t,u,v,w,x,y,z) values ('" \
                +excelfile+"',"+str(r)+",'"+wsheet_display_name
            for c in range(1,27):
                cell_obj=sheet.cell(row=r,column=c)
                if str(cell_obj.value)=='None':
                    cell_val=''
                else:          
                    cell_val=clean_cellval(str(cell_obj.value))

                sql_expr=sql_expr+"','"+cell_val

            sql_expr=sql_expr+"');\n"
            #print(sql_expr)
            f_sql.write(sql_expr)
    
        f_sql.write('commit;\n')
        f_sql.close()

        if view_gen=='Y':
            f_view_sql = open(f_view_name,'w')
            viewname=clean_str(excelfile[0:20])+'_'+wsheet_display_name[0:8]
            viewname=viewname[0:28]+"_v"
            
            sql_expr='create or replace view '+viewname+" ( filename, rn, sheet,\n"
            #process columns A..Z
            for c in range(1,27):
                c_name=clean_str(str(sheet.cell(row=1,column=c).value))
                if c_name=='None':
                    c_name="Column_"+get_column_letter(c)
                c_name=c_name[0:30]
                # Handle some reserved words. Extend if you need more.
                # For those excel cell values the 'Col_' prefix is added, so SQL remains runnable.
                if c_name=="Name" or \
                   c_name=="Comment" or \
                   c_name=="Reference" or \
                   c_name=="Order" or \
                   c_name=="Condition":
                    c_name="Col_"+c_name
                
                if c>1:
                    sql_expr=sql_expr + '   , '+c_name[0:30]+'\n'
                else:
                    sql_expr=sql_expr + '     '+c_name[0:30]+'\n'

            sql_expr=sql_expr+")\n as \nselect filename, rn, sheet\n"

            for c in range(1,27):
                c_letter=get_column_letter(c)
                sql_expr=sql_expr + "   , "+c_letter+"\n"

            sql_expr=sql_expr+"from excel_store\n"
            
            where_cond="where rn>1 and filename='"+excelfile+"'\n" +\
               "order by filename, rn;\n"
                   
            sql_expr=sql_expr+where_cond
            f_view_sql.write(sql_expr)
            f_view_sql.close()

        print("Done. Check file(s): "+f_name+" ,(  *_v.sql)")
    else:
        print("Skipping sheet... "+wsheet)
